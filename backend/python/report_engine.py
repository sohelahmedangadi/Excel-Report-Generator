#!/usr/bin/env python3
"""
DataSheet AI — Report Engine v5
Handles:
  - Symbol stripping (₹ % , $ £ € etc.)
  - Smart numeric detection
  - Missing value analysis
  - Meaningful pivot tables (category-based, not ID-based)
  - 6 embedded charts
  - Insight section
  - Data cleaning report
"""

import sys, os, io, json, warnings, traceback, re
from datetime import datetime

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from matplotlib.patches import FancyBboxPatch

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, CellIsRule

warnings.filterwarnings('ignore')

# ── Palette ───────────────────────────────────────────────────────────────────
DARK_BLUE    = "1A3A5C"
SLATE_BLUE   = "2D5F8A"
MID_BLUE     = "3B82F6"
ACCENT_GREEN = "16A34A"
LIGHT_GREEN  = "DCFCE7"
LIGHT_BLUE   = "EFF6FF"
ORANGE       = "F59E0B"
RED          = "DC2626"
LIGHT_RED    = "FEF2F2"
WHITE        = "FFFFFF"
GRAY_50      = "F8FAFC"
GRAY_100     = "F1F5F9"
GRAY_200     = "E2E8F0"
TEXT_DARK    = "1E293B"
TEXT_MID     = "475569"
TEXT_LIGHT   = "94A3B8"
BORDER_C     = "CBD5E1"

CHART_COLORS = ['#2563EB','#16A34A','#F59E0B','#8B5CF6','#EF4444',
                '#06B6D4','#DB2777','#65A30D','#EA580C','#0D9488']

# ── Progress emitter ──────────────────────────────────────────────────────────
def emit(stage, pct, msg="", meta=None):
    print(json.dumps({"stage":stage,"pct":pct,"msg":msg,**(meta or {})}), flush=True)

# ── Style helpers ─────────────────────────────────────────────────────────────
def _side(c=BORDER_C): return Side(style='thin', color=c)
def _border(c=BORDER_C): s=_side(c); return Border(left=s,right=s,top=s,bottom=s)

def _hdr(ws, r, c, val, bg=SLATE_BLUE, fg=WHITE, sz=10, wrap=True):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font      = Font(bold=True, color=fg, size=sz, name='Calibri')
    cell.fill      = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=wrap)
    cell.border    = _border()
    return cell

def _cell(ws, r, c, val, fmt=None, bold=False, bg=None, fg=TEXT_DARK, align='left', wrap=False):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font      = Font(bold=bold, color=fg, size=10, name='Calibri')
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    cell.border    = _border()
    if bg:  cell.fill = PatternFill('solid', fgColor=bg)
    if fmt: cell.number_format = fmt
    return cell

def _col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

def _merge_title(ws, cell_range, text, bg=DARK_BLUE, fg=WHITE, sz=13, height=None):
    ws.merge_cells(cell_range)
    c = ws[cell_range.split(':')[0]]
    c.value = text
    c.font  = Font(bold=True, size=sz, color=fg, name='Calibri')
    c.fill  = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='center', vertical='center')
    if height:
        row_num = int(''.join(filter(str.isdigit, cell_range.split(':')[0])))
        ws.row_dimensions[row_num].height = height
    return c

def _chart_buf(fig, dpi=150):
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=dpi, bbox_inches='tight',
                facecolor='white', edgecolor='none')
    buf.seek(0)
    plt.close(fig)
    return buf

# ── Data Loading ──────────────────────────────────────────────────────────────
def _sanitize_columns(df):
    """
    Drop fully-empty columns and columns whose name starts with 'Unnamed'.
    Also strip whitespace from column names.
    """
    # Strip column name whitespace
    df.columns = [str(c).strip() for c in df.columns]
    # Drop columns named "Unnamed: N" that are completely empty or all-NaN
    unnamed = [c for c in df.columns if re.match(r'^Unnamed[:\s]', c, re.I)]
    df = df.drop(columns=[c for c in unnamed if df[c].isna().all() or df[c].astype(str).str.strip().eq('').all()])
    # Drop remaining Unnamed columns regardless (they hold no useful header)
    unnamed_left = [c for c in df.columns if re.match(r'^Unnamed[:\s]', c, re.I)]
    df = df.drop(columns=unnamed_left)
    # Drop columns that are entirely NaN
    df = df.dropna(axis=1, how='all')
    return df

def load_data(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == '.csv':
        for enc in ('utf-8', 'utf-8-sig', 'latin-1', 'cp1252'):
            try:
                df = pd.read_csv(path, encoding=enc, low_memory=False)
                return _sanitize_columns(df)
            except UnicodeDecodeError:
                continue
        raise ValueError("Could not decode file — try saving as UTF-8")
    else:
        df = pd.read_excel(path, engine='openpyxl')
        return _sanitize_columns(df)

# ── Smart Data Cleaning ───────────────────────────────────────────────────────
CURRENCY_SYMBOLS = r'[\$£€₹₩¥₺₿₦₫＄]'
CLEAN_PATTERN    = re.compile(r'[\$£€₹₩¥₺₿₦₫＄,%\s]')

def _try_numeric(series):
    """Strip currency/percent symbols and commas, then coerce to numeric."""
    s = series.astype(str).str.strip()
    # Remove currency symbols, commas, percent signs, spaces
    s = s.str.replace(CLEAN_PATTERN, '', regex=True)
    # Handle parentheses for negatives: (1,234) → -1234
    s = s.str.replace(r'^\((.+)\)$', r'-\1', regex=True)
    return pd.to_numeric(s, errors='coerce')

def clean_data(df):
    """
    Full cleaning pipeline:
    1. Strip whitespace
    2. Strip symbols and convert numeric-looking object columns
    3. Parse dates
    4. Drop full duplicates
    Returns df, cleaning_log
    """
    cleaning_log = []
    original_rows = len(df)
    original_dtypes = df.dtypes.copy()

    # 1. Strip whitespace from all string columns
    for col in df.select_dtypes(include='object').columns:
        df[col] = df[col].astype(str).str.strip()
        df[col] = df[col].replace({'nan': np.nan, 'None': np.nan, '': np.nan, 'N/A': np.nan, 'NA': np.nan, '-': np.nan})

    # 2. Smart numeric detection — strip symbols and try to convert
    converted_cols = []
    for col in df.select_dtypes(include='object').columns:
        converted = _try_numeric(df[col])
        valid_ratio = converted.notna().sum() / max(len(df), 1)
        # If 60%+ of non-null values are numeric after stripping, convert
        non_null_mask = df[col].notna()
        non_null_converted = converted[non_null_mask]
        if len(non_null_converted) > 0:
            converted_ratio = non_null_converted.notna().sum() / len(non_null_converted)
        else:
            converted_ratio = 0

        if converted_ratio >= 0.60:
            sample_raw = df[col].dropna().iloc[0] if df[col].notna().any() else ""
            df[col] = converted
            converted_cols.append(col)
            cleaning_log.append({
                "action": "numeric_convert",
                "column": col,
                "sample_raw": str(sample_raw),
                "note": f"Stripped symbols → numeric ({converted_ratio*100:.0f}% valid)"
            })

    # 3. Date parsing for date-named columns
    for col in df.select_dtypes(include='object').columns:
        if any(k in col.lower() for k in ['date','time','created','updated','at']):
            try:
                parsed = pd.to_datetime(df[col], errors='coerce', infer_datetime_format=True)
                if parsed.notna().sum() / max(len(df),1) > 0.5:
                    df[col] = parsed
                    cleaning_log.append({"action":"date_parse","column":col,"note":"Parsed as datetime"})
            except: pass

    # 4. Drop full duplicates
    before = len(df)
    df = df.drop_duplicates()
    removed = before - len(df)
    if removed > 0:
        cleaning_log.append({"action":"dedup","note":f"Removed {removed} duplicate rows"})

    return df, cleaning_log

# ── Column Intelligence ───────────────────────────────────────────────────────
def analyze_columns(df):
    """Return detailed column metadata."""
    info = []
    num_cols = df.select_dtypes(include='number').columns.tolist()
    for col in df.columns:
        null_count = int(df[col].isna().sum())
        non_null   = len(df) - null_count
        dtype_str  = str(df[col].dtype)
        is_num     = col in num_cols

        entry = {
            "name":       col,
            "dtype":      dtype_str,
            "non_null":   non_null,
            "null_count": null_count,
            "null_pct":   f"{df[col].isna().mean()*100:.1f}%",
            "unique":     int(df[col].nunique()),
            "sample":     str(df[col].dropna().iloc[0]) if non_null > 0 else "",
        }
        if is_num and non_null > 0:
            entry["mean"] = round(float(df[col].mean()), 2)
            entry["std"]  = round(float(df[col].std()),  2)
            entry["min"]  = round(float(df[col].min()),  2)
            entry["max"]  = round(float(df[col].max()),  2)
        info.append(entry)
    return info

def pick_cols(df):
    """
    Pick best categorical + numeric pair for pivot/charts.
    Avoids ID columns (high cardinality) for categorical.
    Prefers columns with meaningful aggregation for numeric.
    """
    num_cols = df.select_dtypes(include='number').columns.tolist()
    obj_cols = df.select_dtypes(include='object').columns.tolist()

    # Categorical: prefer 2–50 unique values, penalise ID-like names
    ID_PATTERNS = re.compile(r'\b(id|uuid|key|code|sku|index|no\.?|#)\b', re.I)
    def cat_score(c):
        u = df[c].nunique()
        if u < 2 or u > 100: return 9999
        penalty = 50 if ID_PATTERNS.search(c) else 0
        return u + penalty

    cat_candidates = sorted(obj_cols, key=cat_score)
    cat = cat_candidates[0] if cat_candidates else (df.columns[0])

    # Also check low-cardinality numeric cols (e.g. star ratings used as category)
    # but skip them if we have a proper string cat column

    # Numeric: prefer price/amount/sales/revenue/count columns, exclude IDs
    NUM_PREFER = re.compile(r'\b(price|amount|revenue|sales|cost|value|count|score|rating|total|qty|quantity|avg|mean)\b', re.I)
    def num_score(c):
        if c == cat: return 9999
        if ID_PATTERNS.search(c): return 9998
        prefer = -100 if NUM_PREFER.search(c) else 0
        null_ratio = df[c].isna().mean()
        return null_ratio * 100 + prefer

    num_candidates = sorted([c for c in num_cols if c != cat], key=num_score)

    if num_candidates:
        num = num_candidates[0]
    elif num_cols:
        num = num_cols[0]
    else:
        df['_count'] = 1
        num = '_count'

    return cat, num

def find_col(df, *keywords):
    """Return first column whose name contains any keyword (case-insensitive), or None."""
    for kw in keywords:
        for col in df.columns:
            if kw.lower() in col.lower():
                return col
    return None

def find_num_col(df, *keywords):
    """Like find_col but only returns numeric columns."""
    num_cols = set(df.select_dtypes(include='number').columns)
    for kw in keywords:
        for col in df.columns:
            if kw.lower() in col.lower() and col in num_cols:
                return col
    return None

def detect_fraud_col(df):
    """
    Return (fraud_col, is_binary) if we find a likely fraud/label column.
    is_binary=True  → column contains 0/1 or True/False or 'fraud'/'not fraud'
    is_binary=False → column is categorical (e.g. transaction type)
    """
    FRAUD_KEYWORDS = re.compile(
        r'\b(fraud|is_fraud|isFraud|label|class|target|flag|chargeback|dispute)\b', re.I)
    for col in df.columns:
        if FRAUD_KEYWORDS.search(col):
            u = df[col].dropna().unique()
            # Binary: {0,1} or {True,False} or 2-value string column
            if set(str(v).lower() for v in u).issubset({'0','1','true','false','yes','no','fraud','legitimate','not fraud'}):
                return col, True
            if len(u) == 2:
                return col, True
    return None, False

def detect_outliers(df, col, z_thresh=3.0):
    """
    Return (outlier_df, lower_fence, upper_fence) using IQR method.
    outlier_df contains the rows where col is beyond fences.
    """
    data = df[col].dropna()
    if len(data) < 10:
        return pd.DataFrame(), None, None
    q1, q3 = data.quantile(0.25), data.quantile(0.75)
    iqr = q3 - q1
    lower = q1 - 1.5 * iqr
    upper = q3 + 1.5 * iqr
    outliers = df[(df[col] < lower) | (df[col] > upper)]
    return outliers, round(float(lower), 4), round(float(upper), 4)

# ── Chart generators ──────────────────────────────────────────────────────────
def _style_ax(ax, title):
    ax.set_title(title, fontsize=12, fontweight='bold', color='#1A3A5C', pad=12)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_color('#E2E8F0')
    ax.spines['bottom'].set_color('#E2E8F0')
    ax.tick_params(colors='#475569', labelsize=9)
    ax.set_facecolor('#FAFCFF')
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(
        lambda v,_: f'{v/1e6:.1f}M' if abs(v)>=1e6 else f'{v/1e3:.0f}K' if abs(v)>=1000 else f'{v:,.1f}'
    ))

def chart_bar_top(df, cat, num, title, top_n=10):
    """Horizontal bar: top N categories by numeric value."""
    d = df.groupby(cat)[num].mean().sort_values().tail(top_n)
    fig, ax = plt.subplots(figsize=(9, max(4, len(d)*0.45)))
    colors = [CHART_COLORS[i % len(CHART_COLORS)] for i in range(len(d))]
    bars = ax.barh(d.index.astype(str), d.values, color=colors, edgecolor='white', linewidth=0.8, height=0.65)
    # Value labels
    for bar, val in zip(bars, d.values):
        ax.text(bar.get_width()*1.01, bar.get_y()+bar.get_height()/2,
                f'{val:,.1f}', va='center', ha='left', fontsize=8, color='#475569')
    _style_ax(ax, title)
    ax.set_xlabel(num, fontsize=9, color='#475569')
    fig.tight_layout()
    return _chart_buf(fig)

def chart_bar_count(df, cat, title):
    """Bar chart: products per category."""
    d = df[cat].value_counts().head(12)
    fig, ax = plt.subplots(figsize=(9, 4.5))
    bars = ax.bar(d.index.astype(str), d.values, color=CHART_COLORS[:len(d)],
                  edgecolor='white', linewidth=0.8, width=0.7)
    for bar in bars:
        ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.3,
                str(int(bar.get_height())), ha='center', va='bottom', fontsize=8, color='#475569')
    _style_ax(ax, title)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v,_: f'{int(v):,}'))
    ax.tick_params(axis='x', rotation=35)
    ax.set_ylabel('Count', fontsize=9, color='#475569')
    fig.tight_layout()
    return _chart_buf(fig)

def chart_pie(df, cat, num, title):
    """Donut pie: share of total by category."""
    d = df.groupby(cat)[num].sum().sort_values(ascending=False).head(8)
    d = d[d > 0]
    if d.empty: d = pd.Series([1], index=['No data'])
    fig, ax = plt.subplots(figsize=(8, 5))
    wedges, texts, autotexts = ax.pie(
        d.values, labels=None, autopct='%1.1f%%',
        colors=CHART_COLORS[:len(d)], startangle=140,
        wedgeprops=dict(edgecolor='white', linewidth=2.5, width=0.6),
        pctdistance=0.78
    )
    for at in autotexts: at.set_fontsize(8); at.set_color('white'); at.set_fontweight('bold')
    ax.legend(d.index.astype(str), loc='center left', bbox_to_anchor=(0.92, 0.5),
              fontsize=8, frameon=False)
    ax.set_title(title, fontsize=12, fontweight='bold', color='#1A3A5C', pad=12)
    fig.tight_layout()
    return _chart_buf(fig)

def chart_hist(df, col, title):
    """Histogram of a numeric column."""
    data = df[col].dropna()
    if len(data) < 2: return None
    fig, ax = plt.subplots(figsize=(9, 4))
    n, bins, patches = ax.hist(data, bins=min(30, len(data)//5+5),
                                color='#3B82F6', edgecolor='white', linewidth=0.6, alpha=0.85)
    # Color gradient
    for i, patch in enumerate(patches):
        patch.set_facecolor(CHART_COLORS[int(i/len(patches)*len(CHART_COLORS))])
    # Mean line
    mean_val = data.mean()
    ax.axvline(mean_val, color='#EF4444', linestyle='--', lw=1.8, label=f'Mean: {mean_val:,.1f}')
    ax.legend(fontsize=9, frameon=False)
    _style_ax(ax, title)
    ax.set_xlabel(col, fontsize=9, color='#475569')
    ax.set_ylabel('Frequency', fontsize=9, color='#475569')
    fig.tight_layout()
    return _chart_buf(fig)

def chart_scatter(df, x_col, y_col, title):
    """Scatter plot between two numeric columns."""
    d = df[[x_col, y_col]].dropna()
    if len(d) < 3: return None
    fig, ax = plt.subplots(figsize=(9, 4.5))
    ax.scatter(d[x_col], d[y_col], alpha=0.45, color='#3B82F6', s=18, edgecolors='white', lw=0.4)
    # Trend line
    try:
        z = np.polyfit(d[x_col], d[y_col], 1)
        p = np.poly1d(z)
        xs = np.linspace(d[x_col].min(), d[x_col].max(), 100)
        ax.plot(xs, p(xs), color='#EF4444', lw=1.8, linestyle='--', label='Trend')
        ax.legend(fontsize=9, frameon=False)
    except: pass
    _style_ax(ax, title)
    ax.set_xlabel(x_col, fontsize=9, color='#475569')
    ax.set_ylabel(y_col, fontsize=9, color='#475569')
    fig.tight_layout()
    return _chart_buf(fig)

def chart_box(df, cat, num, title, top_n=8):
    """Box plot: distribution of numeric by category."""
    top_cats = df[cat].value_counts().head(top_n).index.tolist()
    d = df[df[cat].isin(top_cats)]
    groups = [d[d[cat]==c][num].dropna().values for c in top_cats]
    groups = [(g, c) for g, c in zip(groups, top_cats) if len(g) >= 2]
    if not groups: return None
    vals, labels = zip(*groups)
    fig, ax = plt.subplots(figsize=(9, 4.5))
    bp = ax.boxplot(vals, patch_artist=True, notch=False, widths=0.55,
                    medianprops=dict(color='white', linewidth=2))
    for patch, color in zip(bp['boxes'], CHART_COLORS):
        patch.set_facecolor(color); patch.set_alpha(0.8)
    ax.set_xticklabels([str(l)[:18] for l in labels], rotation=30, ha='right', fontsize=8)
    _style_ax(ax, title)
    ax.set_ylabel(num, fontsize=9, color='#475569')
    fig.tight_layout()
    return _chart_buf(fig)

def chart_rating_distribution(df, rating_col, title):
    """Bar chart of rating buckets: <3.5, 3.5-4.0, 4.0-4.5, 4.5-5.0."""
    data = df[rating_col].dropna()
    if len(data) < 2: return None
    bins   = [0, 3.5, 4.0, 4.5, 5.01]
    labels = ['Below 3.5', '3.5 – 4.0', '4.0 – 4.5', '4.5 – 5.0']
    counts = pd.cut(data, bins=bins, labels=labels, right=False).value_counts().reindex(labels, fill_value=0)
    colors = ['#EF4444', '#F59E0B', '#3B82F6', '#16A34A']
    fig, ax = plt.subplots(figsize=(8, 4.5))
    bars = ax.bar(labels, counts.values, color=colors, edgecolor='white', linewidth=0.8, width=0.6)
    for bar in bars:
        h = bar.get_height()
        if h > 0:
            ax.text(bar.get_x()+bar.get_width()/2, h+0.3, f'{int(h):,}',
                    ha='center', va='bottom', fontsize=9, fontweight='bold', color='#1A3A5C')
    total = len(data)
    for bar, cnt in zip(bars, counts.values):
        if cnt > 0:
            pct = cnt / total * 100
            ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()/2, f'{pct:.1f}%',
                    ha='center', va='center', fontsize=8, color='white', fontweight='bold')
    ax.set_title(title, fontsize=12, fontweight='bold', color='#1A3A5C', pad=12)
    ax.set_ylabel('Number of Products', fontsize=9, color='#475569')
    ax.set_xlabel('Rating Range', fontsize=9, color='#475569')
    ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
    ax.spines['left'].set_color('#E2E8F0'); ax.spines['bottom'].set_color('#E2E8F0')
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v,_: f'{int(v):,}'))
    ax.set_facecolor('#FAFCFF')
    fig.tight_layout()
    return _chart_buf(fig)

def chart_discount_distribution(df, discount_col, title):
    """Bar chart of discount buckets: 0-20%, 20-40%, 40-60%, 60%+."""
    data = df[discount_col].dropna()
    if len(data) < 2: return None
    bins   = [0, 20, 40, 60, 101]
    labels = ['0 – 20%', '20 – 40%', '40 – 60%', '60%+']
    counts = pd.cut(data, bins=bins, labels=labels, right=False).value_counts().reindex(labels, fill_value=0)
    colors = ['#06B6D4', '#3B82F6', '#8B5CF6', '#EF4444']
    fig, ax = plt.subplots(figsize=(8, 4.5))
    bars = ax.bar(labels, counts.values, color=colors, edgecolor='white', linewidth=0.8, width=0.6)
    for bar in bars:
        h = bar.get_height()
        if h > 0:
            ax.text(bar.get_x()+bar.get_width()/2, h+0.3, f'{int(h):,}',
                    ha='center', va='bottom', fontsize=9, fontweight='bold', color='#1A3A5C')
    total = len(data)
    for bar, cnt in zip(bars, counts.values):
        if cnt > 0:
            pct = cnt / total * 100
            ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()/2, f'{pct:.1f}%',
                    ha='center', va='center', fontsize=8, color='white', fontweight='bold')
    ax.set_title(title, fontsize=12, fontweight='bold', color='#1A3A5C', pad=12)
    ax.set_ylabel('Number of Products', fontsize=9, color='#475569')
    ax.set_xlabel('Discount Range', fontsize=9, color='#475569')
    ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
    ax.spines['left'].set_color('#E2E8F0'); ax.spines['bottom'].set_color('#E2E8F0')
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v,_: f'{int(v):,}'))
    ax.set_facecolor('#FAFCFF')
    fig.tight_layout()
    return _chart_buf(fig)

def chart_category_products(df, cat_col, title):
    """Horizontal bar: number of products per category, sorted descending."""
    raw = df[cat_col].astype(str).str.split('|').str[-1].str.strip()
    d   = raw.value_counts().head(12).sort_values()
    fig, ax = plt.subplots(figsize=(9, max(4, len(d)*0.5)))
    colors = [CHART_COLORS[i % len(CHART_COLORS)] for i in range(len(d))]
    bars   = ax.barh(d.index, d.values, color=colors, edgecolor='white', linewidth=0.8, height=0.65)
    for bar in bars:
        w = bar.get_width()
        ax.text(w + 0.3, bar.get_y()+bar.get_height()/2, f'{int(w):,}',
                va='center', ha='left', fontsize=8, color='#475569')
    ax.set_title(title, fontsize=12, fontweight='bold', color='#1A3A5C', pad=12)
    ax.set_xlabel('Number of Products', fontsize=9, color='#475569')
    ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
    ax.spines['left'].set_color('#E2E8F0'); ax.spines['bottom'].set_color('#E2E8F0')
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda v,_: f'{int(v):,}'))
    ax.set_facecolor('#FAFCFF')
    fig.tight_layout()
    return _chart_buf(fig)

def chart_fraud_vs_legit(df, fraud_col, title):
    """Grouped bar: count of fraud vs non-fraud."""
    vc = df[fraud_col].value_counts().sort_index()
    labels = [str(v) for v in vc.index]
    counts = vc.values
    colors = ['#16A34A' if str(l).lower() in ('0','false','no','legitimate','not fraud')
              else '#EF4444' for l in labels]
    fig, ax = plt.subplots(figsize=(7, 4))
    bars = ax.bar(labels, counts, color=colors, edgecolor='white', linewidth=0.8, width=0.5)
    total = sum(counts)
    for bar, cnt in zip(bars, counts):
        pct = cnt / total * 100
        ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()+total*0.005,
                f'{cnt:,}\n({pct:.1f}%)', ha='center', va='bottom', fontsize=9,
                fontweight='bold', color='#1A3A5C')
    ax.set_title(title, fontsize=12, fontweight='bold', color='#1A3A5C', pad=12)
    ax.set_xlabel(fraud_col, fontsize=9, color='#475569')
    ax.set_ylabel('Count', fontsize=9, color='#475569')
    ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
    ax.spines['left'].set_color('#E2E8F0'); ax.spines['bottom'].set_color('#E2E8F0')
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v,_: f'{int(v):,}'))
    ax.set_facecolor('#FAFCFF')
    fig.tight_layout()
    return _chart_buf(fig)

def chart_fraud_by_category(df, fraud_col, cat_col, title):
    """Stacked bar: fraud vs non-fraud count per category (top 10 cats)."""
    top_cats = df[cat_col].value_counts().head(10).index.tolist()
    d = df[df[cat_col].isin(top_cats)].copy()
    # Normalise fraud value to 0/1 int
    def to_01(v):
        s = str(v).lower()
        return 1 if s in ('1','true','yes','fraud') else 0
    d['_fraud_int'] = d[fraud_col].apply(to_01)
    grp = d.groupby(cat_col)['_fraud_int'].agg(['sum','count'])
    grp['legit'] = grp['count'] - grp['sum']
    grp = grp.reindex(top_cats).dropna()

    fig, ax = plt.subplots(figsize=(9, max(4, len(grp)*0.55)))
    y = range(len(grp))
    ax.barh(list(y), grp['legit'].values, color='#16A34A', label='Legitimate', height=0.55, edgecolor='white')
    ax.barh(list(y), grp['sum'].values,   color='#EF4444', label='Fraud',      height=0.55,
            left=grp['legit'].values, edgecolor='white')
    ax.set_yticks(list(y))
    ax.set_yticklabels([str(c)[:22] for c in grp.index], fontsize=8)
    ax.legend(fontsize=9, frameon=False, loc='lower right')
    ax.set_title(title, fontsize=12, fontweight='bold', color='#1A3A5C', pad=12)
    ax.set_xlabel('Count', fontsize=9, color='#475569')
    ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
    ax.spines['left'].set_color('#E2E8F0'); ax.spines['bottom'].set_color('#E2E8F0')
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda v,_: f'{int(v):,}'))
    ax.set_facecolor('#FAFCFF')
    fig.tight_layout()
    return _chart_buf(fig)

def chart_amount_by_fraud(df, fraud_col, amount_col, title):
    """Box plot: transaction amount for fraud vs non-fraud."""
    def to_01(v):
        return 1 if str(v).lower() in ('1','true','yes','fraud') else 0
    df2 = df[[fraud_col, amount_col]].dropna().copy()
    df2['_f'] = df2[fraud_col].apply(to_01)
    fraud_vals  = df2[df2['_f']==1][amount_col].values
    legit_vals  = df2[df2['_f']==0][amount_col].values
    if len(fraud_vals) < 2 or len(legit_vals) < 2: return None
    fig, ax = plt.subplots(figsize=(7, 4.5))
    bp = ax.boxplot([legit_vals, fraud_vals], patch_artist=True,
                    labels=['Legitimate','Fraud'],
                    medianprops=dict(color='white', linewidth=2.5))
    for patch, color in zip(bp['boxes'], ['#16A34A','#EF4444']):
        patch.set_facecolor(color); patch.set_alpha(0.8)
    ax.set_title(title, fontsize=12, fontweight='bold', color='#1A3A5C', pad=12)
    ax.set_ylabel(amount_col, fontsize=9, color='#475569')
    ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
    ax.spines['left'].set_color('#E2E8F0'); ax.spines['bottom'].set_color('#E2E8F0')
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(
        lambda v,_: f'${v/1e6:.1f}M' if abs(v)>=1e6 else f'${v/1e3:.0f}K' if abs(v)>=1000 else f'${v:,.0f}'
    ))
    ax.set_facecolor('#FAFCFF')
    fig.tight_layout()
    return _chart_buf(fig)

def chart_outlier_boxplot(df, col, title):
    """Box plot with outlier dots highlighted in red."""
    data = df[col].dropna()
    if len(data) < 5: return None
    fig, ax = plt.subplots(figsize=(7, 4))
    bp = ax.boxplot(data.values, vert=True, patch_artist=True,
                    flierprops=dict(marker='o', markerfacecolor='#EF4444',
                                   markeredgecolor='white', markersize=4, alpha=0.5),
                    medianprops=dict(color='white', linewidth=2.5))
    bp['boxes'][0].set_facecolor('#3B82F6'); bp['boxes'][0].set_alpha(0.8)
    ax.set_title(title, fontsize=12, fontweight='bold', color='#1A3A5C', pad=12)
    ax.set_ylabel(col, fontsize=9, color='#475569')
    ax.set_xticklabels([col[:20]])
    ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
    ax.spines['left'].set_color('#E2E8F0'); ax.spines['bottom'].set_color('#E2E8F0')
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(
        lambda v,_: f'{v/1e6:.1f}M' if abs(v)>=1e6 else f'{v/1e3:.0f}K' if abs(v)>=1000 else f'{v:,.0f}'
    ))
    ax.set_facecolor('#FAFCFF')
    # Annotation
    _, lower, upper = detect_outliers(df, col)
    if lower is not None:
        n_out = int(((data < lower) | (data > upper)).sum())
        ax.text(0.98, 0.98, f'{n_out} outliers detected\n(IQR method)',
                transform=ax.transAxes, ha='right', va='top', fontsize=8,
                color='#EF4444', style='italic')
    fig.tight_layout()
    return _chart_buf(fig)

# ── Sheet: Summary / Cover ────────────────────────────────────────────────────
def build_cover(wb, col_info, stats, cleaning_log, insights):
    ws = wb.active
    ws.title = "📊 Summary"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3

    # ── Hero banner
    ws.merge_cells('B1:K3')
    c = ws['B1']
    c.value = "DATASHEET AI — ANALYSIS REPORT"
    c.font  = Font(bold=True, size=20, color=WHITE, name='Calibri')
    c.fill  = PatternFill('solid', fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal='center', vertical='center')
    for r in [1,2,3]: ws.row_dimensions[r].height = 22

    ws.merge_cells('B4:K4')
    sub = ws['B4']
    sub.value = f"Generated  {datetime.now().strftime('%d %B %Y  %H:%M UTC')}"
    sub.font  = Font(size=9, italic=True, color=TEXT_LIGHT, name='Calibri')
    sub.fill  = PatternFill('solid', fgColor=GRAY_100)
    sub.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[4].height = 16

    # ── KPI cards row
    ws.row_dimensions[6].height = 14
    ws.row_dimensions[7].height = 36
    ws.row_dimensions[8].height = 24
    ws.row_dimensions[9].height = 14

    kpi_defs = [
        ('B','C', str(stats['rows']),         'Total Rows',        MID_BLUE),
        ('D','E', str(stats['cols']),          'Total Columns',     SLATE_BLUE),
        ('F','G', str(stats['numeric_cols']),  'Numeric Columns',   ACCENT_GREEN),
        ('H','I', f"{stats['missing_pct']:.1f}%", 'Missing Data',  ORANGE if stats['missing_pct']>5 else ACCENT_GREEN),
        ('J','K', str(stats['duplicate_rows']),'Duplicate Rows',   RED if stats['duplicate_rows']>0 else ACCENT_GREEN),
    ]
    for start_col, end_col, val, label, color in kpi_defs:
        ws.merge_cells(f'{start_col}7:{end_col}7')
        v = ws[f'{start_col}7']
        v.value = val
        v.font  = Font(bold=True, size=18, color=color, name='Calibri')
        v.fill  = PatternFill('solid', fgColor=GRAY_50)
        v.alignment = Alignment(horizontal='center', vertical='center')
        v.border = _border(GRAY_200)

        ws.merge_cells(f'{start_col}8:{end_col}8')
        l = ws[f'{start_col}8']
        l.value = label
        l.font  = Font(size=8, color=TEXT_MID, name='Calibri')
        l.fill  = PatternFill('solid', fgColor=GRAY_50)
        l.alignment = Alignment(horizontal='center', vertical='center')
        l.border = _border(GRAY_200)

    # ── Data Quality section
    row = 11
    _merge_title(ws, f'B{row}:K{row}', '🔍  DATA QUALITY & CLEANING REPORT', bg=SLATE_BLUE, height=22)
    row += 1
    for ci, h in enumerate(['Action','Column / Detail','Finding'],1):
        _hdr(ws, row, ci+1, h, bg=DARK_BLUE, sz=9)
    row += 1
    for item in cleaning_log:
        bg = LIGHT_GREEN if item['action'] in ('numeric_convert','dedup','date_parse') else LIGHT_BLUE
        _cell(ws, row, 2, item['action'].replace('_',' ').title(), bg=bg, bold=True)
        _cell(ws, row, 3, item.get('column', '—'), bg=bg)
        _cell(ws, row, 4, item.get('note', item.get('sample_raw','')), bg=bg, wrap=True)
        ws.row_dimensions[row].height = 16
        row += 1
    if not cleaning_log:
        ws.merge_cells(f'B{row}:K{row}')
        ws[f'B{row}'].value = 'No cleaning actions needed — data appears clean.'
        ws[f'B{row}'].font = Font(italic=True, color=TEXT_LIGHT, name='Calibri')
        row += 1
    row += 1

    # ── Column overview
    _merge_title(ws, f'B{row}:K{row}', '📋  COLUMN OVERVIEW', bg=SLATE_BLUE, height=22)
    row += 1
    hdrs = ['Column Name','Data Type','Non-Null','Null Count','Null %','Unique','Min','Max','Mean','Sample Value']
    for ci, h in enumerate(hdrs, 2):
        _hdr(ws, row, ci, h, bg=DARK_BLUE, sz=9)
    row += 1
    for i, col in enumerate(col_info):
        bg = LIGHT_BLUE if i % 2 == 0 else WHITE
        is_num = 'int' in col['dtype'] or 'float' in col['dtype']
        type_color = ACCENT_GREEN if is_num else SLATE_BLUE
        _cell(ws, row, 2,  col['name'],      bg=bg, bold=True)
        _cell(ws, row, 3,  col['dtype'],     bg=bg, fg=type_color, bold=is_num)
        _cell(ws, row, 4,  col['non_null'],  bg=bg, align='right', fmt='#,##0')
        _cell(ws, row, 5,  col['null_count'],bg=bg, align='right',
              fg=RED if col['null_count']>0 else TEXT_DARK, fmt='#,##0')
        _cell(ws, row, 6,  col['null_pct'],  bg=bg, align='right')
        _cell(ws, row, 7,  col['unique'],    bg=bg, align='right', fmt='#,##0')
        _cell(ws, row, 8,  col.get('min',''), bg=bg, align='right', fmt='#,##0.##')
        _cell(ws, row, 9,  col.get('max',''), bg=bg, align='right', fmt='#,##0.##')
        _cell(ws, row, 10, col.get('mean',''),bg=bg, align='right', fmt='#,##0.##')
        _cell(ws, row, 11, str(col['sample'])[:40], bg=bg)
        ws.row_dimensions[row].height = 16
        row += 1

    row += 1
    # ── Insights
    if insights:
        _merge_title(ws, f'B{row}:K{row}', '💡  KEY INSIGHTS', bg=ACCENT_GREEN, height=22)
        row += 1
        for ins in insights:
            ws.merge_cells(f'B{row}:K{row}')
            c = ws[f'B{row}']
            c.value = f"  {ins}"
            c.font  = Font(size=10, color=TEXT_DARK, name='Calibri')
            c.fill  = PatternFill('solid', fgColor=LIGHT_GREEN)
            c.border = _border(GRAY_200)
            c.alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[row].height = 18
            row += 1

    _col_widths(ws, {2:22,3:13,4:10,5:10,6:8,7:9,8:11,9:11,10:10,11:28})

# ── Sheet: Cleaned Data ───────────────────────────────────────────────────────
def build_raw(wb, df, cleaning_log=None):
    ws = wb.create_sheet("📄 Cleaned Data")
    ws.sheet_view.showGridLines = False

    n_cols = len(df.columns)
    col_letter = get_column_letter(max(n_cols, 1))

    _merge_title(ws, f'A1:{col_letter}1', 'CLEANED & PROCESSED DATA', height=24)

    # ── Cleaning summary banner (row 2)
    summary_parts = []
    if cleaning_log:
        converts = [l for l in cleaning_log if l['action'] == 'numeric_convert']
        dates    = [l for l in cleaning_log if l['action'] == 'date_parse']
        dedup    = next((l for l in cleaning_log if l['action'] == 'dedup'), None)
        if converts:
            cols_str = ', '.join(f"'{l['column']}'" for l in converts[:4])
            summary_parts.append(f"• {len(converts)} column(s) converted to numeric: {cols_str}")
        if dates:
            summary_parts.append(f"• {len(dates)} column(s) parsed as datetime")
        if dedup:
            summary_parts.append(f"• {dedup['note']}")
        summary_parts.append(f"• Unnamed/empty columns removed  •  Whitespace stripped  •  Null placeholders standardised")
    else:
        summary_parts = ['• No transformations needed — data was already clean.']

    summary_text = '  |  '.join(summary_parts)
    ws.merge_cells(f'A2:{col_letter}2')
    banner = ws['A2']
    banner.value = f'ℹ️  TRANSFORMATIONS APPLIED:  {summary_text}'
    banner.font  = Font(size=9, italic=True, color='1E3A5F', name='Calibri')
    banner.fill  = PatternFill('solid', fgColor='DBEAFE')
    banner.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    banner.border = _border('BFDBFE')
    ws.row_dimensions[2].height = 28

    # ── Header row
    for ci, col in enumerate(df.columns, 1):
        _hdr(ws, 3, ci, col, bg=SLATE_BLUE, sz=9)

    BATCH = 2000
    for start in range(0, len(df), BATCH):
        chunk = df.iloc[start:start+BATCH]
        for ri, (_, row) in enumerate(chunk.iterrows(), start+4):
            bg = GRAY_50 if ri % 2 == 0 else WHITE
            for ci, val in enumerate(row, 1):
                v = None if pd.isna(val) else val
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.fill   = PatternFill('solid', fgColor=bg)
                cell.font   = Font(size=9, color=TEXT_DARK, name='Calibri')
                cell.border = _border()
                if isinstance(v, float): cell.number_format = '#,##0.##'

    # Auto-widths (capped)
    for ci, col in enumerate(df.columns, 1):
        max_len = max(len(str(col)), df[col].astype(str).str.len().max() if len(df) > 0 else 0)
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len+3, 35)
    ws.freeze_panes = 'A4'

# ── Sheet: Statistics ─────────────────────────────────────────────────────────
def build_stats(wb, df):
    ws = wb.create_sheet("📈 Statistics")
    ws.sheet_view.showGridLines = False
    num_df = df.select_dtypes(include='number')
    if num_df.empty:
        ws['A1'] = 'No numeric columns found.'
        return

    _merge_title(ws, f'A1:{get_column_letter(len(num_df.columns)+1)}1',
                 'DESCRIPTIVE STATISTICS — NUMERIC COLUMNS', height=26)

    stats_labels = ['Count','Mean','Std Dev','Min','25%','Median','75%','Max','Sum','Missing','Missing %']
    _hdr(ws, 2, 1, 'Metric', bg=DARK_BLUE, sz=10)
    for ci, col in enumerate(num_df.columns, 2):
        _hdr(ws, 2, ci, col, bg=DARK_BLUE, sz=9)

    desc = num_df.describe(percentiles=[.25,.5,.75])
    stats_rows = {
        'Count':     lambda c: int(num_df[c].count()),
        'Mean':      lambda c: round(float(num_df[c].mean()), 4) if num_df[c].count()>0 else None,
        'Std Dev':   lambda c: round(float(num_df[c].std()), 4) if num_df[c].count()>1 else None,
        'Min':       lambda c: round(float(num_df[c].min()), 4) if num_df[c].count()>0 else None,
        '25%':       lambda c: round(float(num_df[c].quantile(.25)), 4) if num_df[c].count()>0 else None,
        'Median':    lambda c: round(float(num_df[c].median()), 4) if num_df[c].count()>0 else None,
        '75%':       lambda c: round(float(num_df[c].quantile(.75)), 4) if num_df[c].count()>0 else None,
        'Max':       lambda c: round(float(num_df[c].max()), 4) if num_df[c].count()>0 else None,
        'Sum':       lambda c: round(float(num_df[c].sum()), 2) if num_df[c].count()>0 else None,
        'Missing':   lambda c: int(num_df[c].isna().sum()),
        'Missing %': lambda c: f"{num_df[c].isna().mean()*100:.1f}%",
    }

    for ri, (label, fn) in enumerate(stats_rows.items(), 3):
        bg = LIGHT_BLUE if ri % 2 == 0 else WHITE
        _cell(ws, ri, 1, label, bold=True, bg=bg)
        for ci, col in enumerate(num_df.columns, 2):
            val = fn(col)
            cell = _cell(ws, ri, ci, val, bg=bg, align='right')
            if isinstance(val, float):
                cell.number_format = '#,##0.00'

    # Color scale on means row
    mean_row = 4
    col_range = f'B{mean_row}:{get_column_letter(len(num_df.columns)+1)}{mean_row}'
    ws.conditional_formatting.add(col_range, ColorScaleRule(
        start_type='min', start_color='FEF2F2',
        mid_type='percentile', mid_value=50, mid_color='FFFBEB',
        end_type='max', end_color='F0FDF4'
    ))

    _col_widths(ws, {1: 14, **{i+2: 18 for i in range(len(num_df.columns))}})
    ws.freeze_panes = 'B3'

# ── Sheet: Missing Value Analysis ─────────────────────────────────────────────
def build_missing(wb, df):
    ws = wb.create_sheet("🔍 Missing Values")
    ws.sheet_view.showGridLines = False

    _merge_title(ws, f'A1:F1', 'MISSING VALUE ANALYSIS', height=26)
    hdrs = ['Column', 'Total Rows', 'Non-Null', 'Missing Count', 'Missing %', 'Status']
    for ci, h in enumerate(hdrs, 1):
        _hdr(ws, 2, ci, h, bg=DARK_BLUE)

    total = len(df)
    for ri, col in enumerate(df.columns, 3):
        missing = int(df[col].isna().sum())
        pct     = df[col].isna().mean() * 100
        non_null = total - missing
        bg = LIGHT_RED if pct > 20 else (LIGHT_GREEN if pct == 0 else LIGHT_BLUE)

        if pct == 0:   status = '✅ Complete'
        elif pct < 5:  status = '⚠️ Minor gaps'
        elif pct < 20: status = '🟡 Moderate'
        else:          status = '🔴 High missing'

        _cell(ws, ri, 1, col,      bold=True, bg=bg)
        _cell(ws, ri, 2, total,    bg=bg, align='right', fmt='#,##0')
        _cell(ws, ri, 3, non_null, bg=bg, align='right', fmt='#,##0')
        _cell(ws, ri, 4, missing,  bg=bg, align='right', fmt='#,##0',
              fg=RED if missing > 0 else ACCENT_GREEN)
        _cell(ws, ri, 5, round(pct,1), bg=bg, align='right', fmt='0.0%' if False else '0.00')
        _cell(ws, ri, 6, status,   bg=bg)
        ws.row_dimensions[ri].height = 16

    # Data bar on missing count
    last_row = 2 + len(df.columns)
    ws.conditional_formatting.add(f'D3:D{last_row}', DataBarRule(
        start_type='num', start_value=0,
        end_type='max', color='DC2626'
    ))
    _col_widths(ws, {1:25, 2:12, 3:12, 4:14, 5:12, 6:18})

# ── Sheet: Category Analysis ─────────────────────────────────────────────────
def build_category_analysis(wb, df, cat_col, rating_col, discount_col, price_col):
    """
    Dedicated category sheet:
    - Average rating by category
    - Average discount by category
    - Product count by category
    """
    ws = wb.create_sheet("📂 Category Analysis")
    ws.sheet_view.showGridLines = False

    _merge_title(ws, 'A1:H1', 'CATEGORY ANALYSIS', sz=14, height=28)

    # ── Build the aggregation dict dynamically
    agg = {'product_name_count': (cat_col, 'count')}  # placeholder; we'll do it manually

    # Strip sub-category (handle "Electronics|Headphones" → "Electronics|Headphones" kept,
    # but also compute a top-level "main_category")
    df2 = df.copy()
    df2['_cat'] = df2[cat_col].astype(str).str.strip()

    rows_out = []
    for cat_val, grp in df2.groupby('_cat', sort=False):
        row = {'Category': str(cat_val), 'Products': int(len(grp))}
        if rating_col and rating_col in grp.columns:
            row['Avg Rating']    = round(float(grp[rating_col].mean()), 2) if grp[rating_col].notna().any() else None
            row['Min Rating']    = round(float(grp[rating_col].min()), 2)  if grp[rating_col].notna().any() else None
            row['Max Rating']    = round(float(grp[rating_col].max()), 2)  if grp[rating_col].notna().any() else None
        if discount_col and discount_col in grp.columns:
            row['Avg Discount %'] = round(float(grp[discount_col].mean()), 1) if grp[discount_col].notna().any() else None
        if price_col and price_col in grp.columns:
            row['Avg Price']     = round(float(grp[price_col].mean()), 2) if grp[price_col].notna().any() else None
        rows_out.append(row)

    result = pd.DataFrame(rows_out).sort_values('Products', ascending=False)

    # ── Table 1: Full category summary
    row = 3
    sub_title_cell = ws.cell(row=row, column=1)
    ws.merge_cells(f'A{row}:H{row}')
    sub_title_cell.value = f'📊  Summary by {cat_col}  ({len(result)} categories)'
    sub_title_cell.font  = Font(bold=True, size=11, color=WHITE, name='Calibri')
    sub_title_cell.fill  = PatternFill('solid', fgColor=SLATE_BLUE)
    sub_title_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 20
    row += 1

    hdrs = list(result.columns)
    for ci, h in enumerate(hdrs, 1):
        _hdr(ws, row, ci, h, bg=DARK_BLUE, sz=10)
    row += 1

    # Medal colors for top 3
    medals = [LIGHT_GREEN, LIGHT_BLUE, 'FFF9C4', WHITE]
    for ri, (_, drow) in enumerate(result.iterrows()):
        bg = medals[min(ri, 3)]
        bold = ri < 3
        for ci, (col_name, val) in enumerate(drow.items(), 1):
            if pd.isna(val): val = None
            cell = _cell(ws, row, ci, val, bold=bold, bg=bg)
            if isinstance(val, float):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif isinstance(val, int) and col_name == 'Products':
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right', vertical='center')
        ws.row_dimensions[row].height = 16
        row += 1

    # Data bar on Products count
    if len(result) > 1:
        ws.conditional_formatting.add(
            f'B{row - len(result)}:B{row - 1}',
            DataBarRule(start_type='min', end_type='max', color=MID_BLUE)
        )
    row += 2

    # ── Table 2: Rating highlights (if rating col exists)
    if rating_col and rating_col in df.columns:
        ws.merge_cells(f'A{row}:H{row}')
        c2 = ws[f'A{row}']
        c2.value = f'⭐  TOP 10 CATEGORIES BY AVERAGE RATING'
        c2.font  = Font(bold=True, size=11, color=WHITE, name='Calibri')
        c2.fill  = PatternFill('solid', fgColor=ACCENT_GREEN)
        c2.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 20
        row += 1

        top_rated = result.dropna(subset=['Avg Rating']).nlargest(10, 'Avg Rating')[
            ['Category', 'Avg Rating', 'Products']
        ] if 'Avg Rating' in result.columns else pd.DataFrame()

        if not top_rated.empty:
            for ci, h in enumerate(['Category', 'Avg Rating', 'Products'], 1):
                _hdr(ws, row, ci, h, bg=DARK_BLUE, sz=10)
            row += 1
            for ri, (_, drow) in enumerate(top_rated.iterrows()):
                bg = LIGHT_GREEN if ri < 3 else (LIGHT_BLUE if ri % 2 == 0 else WHITE)
                _cell(ws, row, 1, drow['Category'], bold=(ri < 3), bg=bg)
                c = _cell(ws, row, 2, drow['Avg Rating'], bg=bg, align='right')
                c.number_format = '0.00'
                # Star rating visual
                stars = '★' * int(round(drow['Avg Rating'])) + '☆' * (5 - int(round(drow['Avg Rating'])))
                ws.cell(row=row, column=2).value = f"{drow['Avg Rating']:.2f}  {stars}"
                _cell(ws, row, 3, int(drow['Products']), bg=bg, align='right')
                ws.row_dimensions[row].height = 16
                row += 1
        row += 2

    # ── Table 3: Discount highlights (if discount col exists)
    if discount_col and discount_col in df.columns and 'Avg Discount %' in result.columns:
        ws.merge_cells(f'A{row}:H{row}')
        c3 = ws[f'A{row}']
        c3.value = f'💸  TOP 10 CATEGORIES BY AVERAGE DISCOUNT'
        c3.font  = Font(bold=True, size=11, color=WHITE, name='Calibri')
        c3.fill  = PatternFill('solid', fgColor=ORANGE)
        c3.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 20
        row += 1

        top_disc = result.dropna(subset=['Avg Discount %']).nlargest(10, 'Avg Discount %')[
            ['Category', 'Avg Discount %', 'Products']
        ]
        if not top_disc.empty:
            for ci, h in enumerate(['Category', 'Avg Discount %', 'Products'], 1):
                _hdr(ws, row, ci, h, bg=DARK_BLUE, sz=10)
            row += 1
            for ri, (_, drow) in enumerate(top_disc.iterrows()):
                bg = 'FFF3CD' if ri < 3 else (LIGHT_BLUE if ri % 2 == 0 else WHITE)
                _cell(ws, row, 1, drow['Category'], bold=(ri < 3), bg=bg)
                c = _cell(ws, row, 2, drow['Avg Discount %'], bg=bg, align='right')
                c.number_format = '0.0"%"'
                _cell(ws, row, 3, int(drow['Products']), bg=bg, align='right')
                ws.row_dimensions[row].height = 16
                row += 1

    col_widths = {1: 40, 2: 14, 3: 12, 4: 12, 5: 12, 6: 16, 7: 14}
    for ci, w in col_widths.items():
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A5'


# ── Sheet: Top Rated & Most Reviewed Products ─────────────────────────────────
def build_top_products(wb, df, name_col, rating_col, rating_count_col, discount_col, cat_col):
    """
    Sheet 1 section: Top 30 products by rating
    Sheet 2 section: Top 30 most reviewed products
    """
    ws = wb.create_sheet("🏆 Top Products")
    ws.sheet_view.showGridLines = False

    _merge_title(ws, 'A1:G1', 'TOP RATED & MOST REVIEWED PRODUCTS', sz=13, height=28)
    row = 3

    # ── Table 1: Top Rated
    ws.merge_cells(f'A{row}:G{row}')
    t1 = ws[f'A{row}']
    t1.value = '⭐  TOP 30 PRODUCTS BY RATING  (sorted highest → lowest)'
    t1.font  = Font(bold=True, size=11, color=WHITE, name='Calibri')
    t1.fill  = PatternFill('solid', fgColor=ACCENT_GREEN)
    t1.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 20
    row += 1

    # Build columns for this table — deduplicate while preserving order
    seen = set()
    top_r_cols = []
    for c in [name_col, cat_col, rating_col, rating_count_col, discount_col]:
        if c and c not in seen and c in df.columns:
            top_r_cols.append(c)
            seen.add(c)

    if rating_col and rating_col in df.columns and top_r_cols:
        sort_by = [rating_col]
        if rating_count_col and rating_count_col in top_r_cols:
            sort_by.append(rating_count_col)
        top_rated_df = df[top_r_cols].dropna(subset=[rating_col]).sort_values(
            by=sort_by, ascending=False
        ).head(30).reset_index(drop=True)

        display_hdrs = {'product_name':'Product Name','product_title':'Product Name',
                        'name':'Name','title':'Title',
                        'category':'Category','main_category':'Category',
                        'rating':'Rating','ratings':'Rating',
                        'rating_count':'Reviews','no_of_ratings':'Reviews',
                        'discount_percentage':'Discount %'}

        for ci, col in enumerate(top_r_cols, 1):
            h = display_hdrs.get(col.lower(), col.replace('_',' ').title())
            _hdr(ws, row, ci, h, bg=DARK_BLUE, sz=9)
        row += 1

        for ri, (_, drow) in enumerate(top_rated_df.iterrows()):
            bg = LIGHT_GREEN if ri == 0 else ('FFF9C4' if ri < 3 else (LIGHT_BLUE if ri % 2 == 0 else WHITE))
            bold = ri < 3
            for ci, col in enumerate(top_r_cols, 1):
                val = drow[col]
                if pd.isna(val): val = None
                cell = _cell(ws, row, ci, val, bold=bold, bg=bg)
                if col == rating_col and val is not None:
                    cell.number_format = '0.0'
                    stars = '★' * int(round(float(val))) + '☆' * (5 - int(round(float(val))))
                    cell.value = f"{float(val):.1f} {stars}"
                elif isinstance(val, float):
                    cell.number_format = '#,##0.##'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif isinstance(val, (int, np.integer)):
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
            ws.row_dimensions[row].height = 15
            row += 1

    row += 2

    # ── Table 2: Most Reviewed
    ws.merge_cells(f'A{row}:G{row}')
    t2 = ws[f'A{row}']
    t2.value = '📊  TOP 30 MOST REVIEWED PRODUCTS  (sorted by review count)'
    t2.font  = Font(bold=True, size=11, color=WHITE, name='Calibri')
    t2.fill  = PatternFill('solid', fgColor=SLATE_BLUE)
    t2.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 20
    row += 1

    if rating_count_col and rating_count_col in df.columns:
        top_rev_cols = [c for c in top_r_cols if c in df.columns]
        top_rev_df = df[top_rev_cols].dropna(subset=[rating_count_col]).sort_values(
            rating_count_col, ascending=False
        ).head(30).reset_index(drop=True)

        for ci, col in enumerate(top_rev_cols, 1):
            h = display_hdrs.get(col.lower(), col.replace('_',' ').title())
            _hdr(ws, row, ci, h, bg=DARK_BLUE, sz=9)
        row += 1

        for ri, (_, drow) in enumerate(top_rev_df.iterrows()):
            bg = LIGHT_BLUE if ri == 0 else (LIGHT_GREEN if ri < 3 else (GRAY_50 if ri % 2 == 0 else WHITE))
            bold = ri < 3
            for ci, col in enumerate(top_rev_cols, 1):
                val = drow[col]
                if pd.isna(val): val = None
                cell = _cell(ws, row, ci, val, bold=bold, bg=bg)
                if col == rating_col and val is not None:
                    cell.value = f"{float(val):.1f} {'★' * int(round(float(val)))}"
                elif isinstance(val, float):
                    cell.number_format = '#,##0.##'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif isinstance(val, (int, np.integer)):
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
            ws.row_dimensions[row].height = 15
            row += 1

        # Data bar on review count column
        rc_idx = top_rev_cols.index(rating_count_col) + 1 if rating_count_col in top_rev_cols else None
        if rc_idx:
            data_start = row - len(top_rev_df)
            ws.conditional_formatting.add(
                f'{get_column_letter(rc_idx)}{data_start}:{get_column_letter(rc_idx)}{row-1}',
                DataBarRule(start_type='min', end_type='max', color=MID_BLUE)
            )

    # Column widths
    ws.column_dimensions['A'].width = 45  # product name
    for ci in range(2, 8):
        ws.column_dimensions[get_column_letter(ci)].width = 16
    ws.freeze_panes = 'A5'


# ── Sheet: Fraud Pattern Analysis ────────────────────────────────────────────
def build_fraud_analysis(wb, df, fraud_col, cat_col, amount_col, tx_type_col):
    """
    Dedicated fraud sheet — only added when a fraud/label column is found.
    Sections:
      1. Fraud summary KPIs
      2. Fraud by category (pivot)
      3. Fraud by transaction type (pivot)
      4. High-risk breakdown (top categories by fraud rate)
    """
    ws = wb.create_sheet("🚨 Fraud Analysis")
    ws.sheet_view.showGridLines = False
    _merge_title(ws, 'A1:K1', '🚨  FRAUD PATTERN ANALYSIS', bg=RED, sz=14, height=28)

    # Normalise fraud flag → 0/1 numeric column
    def to_01(v):
        return 1 if str(v).lower() in ('1','true','yes','fraud') else 0
    df2 = df.copy()
    df2['_fraud'] = df2[fraud_col].apply(to_01)

    total     = len(df2)
    n_fraud   = int(df2['_fraud'].sum())
    n_legit   = total - n_fraud
    fraud_pct = n_fraud / total * 100 if total > 0 else 0

    # ── KPI row
    row = 3
    ws.row_dimensions[row].height = 36
    ws.row_dimensions[row+1].height = 20

    kpi_data = [
        ('A', 'B', f'{total:,}',       'Total Transactions', MID_BLUE),
        ('C', 'D', f'{n_fraud:,}',     'Fraud Transactions', RED),
        ('E', 'F', f'{n_legit:,}',     'Legitimate',         ACCENT_GREEN),
        ('G', 'H', f'{fraud_pct:.2f}%','Fraud Rate',         ORANGE if fraud_pct < 10 else RED),
    ]
    if amount_col and amount_col in df2.columns:
        avg_fraud_amt = df2[df2['_fraud']==1][amount_col].mean()
        avg_legit_amt = df2[df2['_fraud']==0][amount_col].mean()
        kpi_data.append(('I','J', f'${avg_fraud_amt:,.0f}' if pd.notna(avg_fraud_amt) else '—',
                         'Avg Fraud Amount', RED))

    for sc, ec, val, label, color in kpi_data:
        ws.merge_cells(f'{sc}{row}:{ec}{row}')
        v = ws[f'{sc}{row}']
        v.value = val
        v.font  = Font(bold=True, size=18, color=color, name='Calibri')
        v.fill  = PatternFill('solid', fgColor=GRAY_50)
        v.alignment = Alignment(horizontal='center', vertical='center')
        v.border = _border(GRAY_200)
        ws.merge_cells(f'{sc}{row+1}:{ec}{row+1}')
        l = ws[f'{sc}{row+1}']
        l.value = label
        l.font  = Font(size=8, color=TEXT_MID, name='Calibri')
        l.fill  = PatternFill('solid', fgColor=GRAY_50)
        l.alignment = Alignment(horizontal='center', vertical='center')
        l.border = _border(GRAY_200)
    row += 3

    # ── Section: Fraud by Category
    if cat_col and cat_col in df2.columns:
        ws.merge_cells(f'A{row}:K{row}')
        sec = ws[f'A{row}']
        sec.value = f'📂  FRAUD BY {cat_col.upper()}'
        sec.font  = Font(bold=True, size=11, color=WHITE, name='Calibri')
        sec.fill  = PatternFill('solid', fgColor=SLATE_BLUE)
        sec.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 20
        row += 1

        cat_grp = df2.groupby(cat_col)['_fraud'].agg(
            Total='count', Fraud='sum'
        ).reset_index()
        cat_grp['Legitimate'] = cat_grp['Total'] - cat_grp['Fraud']
        cat_grp['Fraud Rate %'] = (cat_grp['Fraud'] / cat_grp['Total'] * 100).round(2)
        if amount_col and amount_col in df2.columns:
            cat_grp['Avg Amount'] = df2.groupby(cat_col)[amount_col].mean().round(2).values
        cat_grp = cat_grp.sort_values('Fraud Rate %', ascending=False)

        hdrs_cat = list(cat_grp.columns)
        for ci, h in enumerate(hdrs_cat, 1):
            _hdr(ws, row, ci, h, bg=DARK_BLUE, sz=9)
        row += 1
        for ri, (_, r_data) in enumerate(cat_grp.iterrows()):
            bg = LIGHT_RED if r_data['Fraud Rate %'] > fraud_pct * 1.5 else (
                 LIGHT_GREEN if r_data['Fraud Rate %'] < fraud_pct * 0.5 else (
                 LIGHT_BLUE if ri % 2 == 0 else WHITE))
            for ci, (col_name, val) in enumerate(r_data.items(), 1):
                cell = _cell(ws, row, ci, val if not pd.isna(val) else None, bg=bg)
                if isinstance(val, float):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif isinstance(val, (int, np.integer)):
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
            ws.row_dimensions[row].height = 15
            row += 1
        row += 2

    # ── Section: Fraud by Transaction Type
    if tx_type_col and tx_type_col in df2.columns:
        ws.merge_cells(f'A{row}:K{row}')
        sec2 = ws[f'A{row}']
        sec2.value = f'🔄  FRAUD BY {tx_type_col.upper()}'
        sec2.font  = Font(bold=True, size=11, color=WHITE, name='Calibri')
        sec2.fill  = PatternFill('solid', fgColor=ORANGE)
        sec2.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 20
        row += 1

        type_grp = df2.groupby(tx_type_col)['_fraud'].agg(
            Total='count', Fraud='sum'
        ).reset_index()
        type_grp['Legitimate'] = type_grp['Total'] - type_grp['Fraud']
        type_grp['Fraud Rate %'] = (type_grp['Fraud'] / type_grp['Total'] * 100).round(2)
        type_grp = type_grp.sort_values('Fraud Rate %', ascending=False)

        for ci, h in enumerate(list(type_grp.columns), 1):
            _hdr(ws, row, ci, h, bg=DARK_BLUE, sz=9)
        row += 1
        for ri, (_, r_data) in enumerate(type_grp.iterrows()):
            bg = LIGHT_RED if ri == 0 else (LIGHT_BLUE if ri % 2 == 0 else WHITE)
            for ci, val in enumerate(r_data, 1):
                cell = _cell(ws, row, ci, val if not pd.isna(val) else None, bg=bg)
                if isinstance(val, float):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif isinstance(val, (int, np.integer)):
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
            ws.row_dimensions[row].height = 15
            row += 1

    for i in range(1, 12):
        ws.column_dimensions[get_column_letter(i)].width = 16
    ws.column_dimensions['A'].width = 30
    ws.freeze_panes = 'A4'


# ── Sheet: Outlier Detection ──────────────────────────────────────────────────
def build_outlier_analysis(wb, df, num_cols):
    """
    One table per numeric column:
    - IQR fences
    - Outlier count + pct
    - Top 10 outlier rows
    """
    ws = wb.create_sheet("📐 Outlier Detection")
    ws.sheet_view.showGridLines = False
    _merge_title(ws, 'A1:H1', '📐  OUTLIER DETECTION  (IQR Method)', sz=13, height=26)

    row = 3
    for col in num_cols:
        outliers, lower, upper = detect_outliers(df, col)
        n_out = len(outliers)
        n_total = df[col].dropna().count()
        pct = n_out / n_total * 100 if n_total > 0 else 0

        # Section header
        ws.merge_cells(f'A{row}:H{row}')
        sec = ws[f'A{row}']
        sec.value = f'📊  {col}'
        sec.font  = Font(bold=True, size=11, color=WHITE, name='Calibri')
        sec.fill  = PatternFill('solid', fgColor=SLATE_BLUE if n_out == 0 else RED)
        sec.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 20
        row += 1

        # Stats row
        stat_labels = ['Q1 (25th pct)','Q3 (75th pct)','IQR','Lower Fence','Upper Fence','Outlier Count','Outlier %']
        data = df[col].dropna()
        q1 = float(data.quantile(0.25)) if len(data) > 0 else 0
        q3 = float(data.quantile(0.75)) if len(data) > 0 else 0
        iqr_val = q3 - q1
        stat_values = [
            round(q1, 4), round(q3, 4), round(iqr_val, 4),
            lower if lower is not None else 'N/A',
            upper if upper is not None else 'N/A',
            n_out, f'{pct:.1f}%'
        ]
        for ci, (lbl, val) in enumerate(zip(stat_labels, stat_values), 1):
            _hdr(ws, row, ci, lbl, bg=DARK_BLUE, sz=8)
        row += 1
        for ci, val in enumerate(stat_values, 1):
            cell = _cell(ws, row, ci, val,
                         bg=LIGHT_RED if ci in (6,7) and n_out > 0 else LIGHT_BLUE,
                         align='right')
            if isinstance(val, float): cell.number_format = '#,##0.00'
            elif isinstance(val, int): cell.number_format = '#,##0'
        ws.row_dimensions[row].height = 16
        row += 1

        # Show top outlier rows (up to 10)
        if n_out > 0 and not outliers.empty:
            show_cols = [col] + [c for c in df.columns if c != col][:4]
            show_cols = [c for c in show_cols if c in outliers.columns]
            top_out = outliers[show_cols].head(10).sort_values(col, ascending=False)

            ws.merge_cells(f'A{row}:H{row}')
            sub = ws[f'A{row}']
            sub.value = f'  Top outlier rows (max {min(10,n_out)} of {n_out:,})'
            sub.font  = Font(italic=True, size=9, color=TEXT_MID, name='Calibri')
            sub.fill  = PatternFill('solid', fgColor=LIGHT_RED)
            row += 1
            for ci, h in enumerate(show_cols, 1):
                _hdr(ws, row, ci, h, bg=DARK_BLUE, sz=8)
            row += 1
            for _, out_row in top_out.iterrows():
                for ci, c in enumerate(show_cols, 1):
                    val = out_row[c]
                    cell = _cell(ws, row, ci, val if not pd.isna(val) else None, bg=LIGHT_RED)
                    if isinstance(val, float): cell.number_format = '#,##0.##'
                ws.row_dimensions[row].height = 14
                row += 1
        else:
            ws.merge_cells(f'A{row}:H{row}')
            ok = ws[f'A{row}']
            ok.value = '  ✅ No outliers detected in this column.'
            ok.font  = Font(italic=True, size=9, color=ACCENT_GREEN, name='Calibri')
            ok.fill  = PatternFill('solid', fgColor=LIGHT_GREEN)
            row += 1

        row += 1  # gap between columns

    for i in range(1, 9):
        ws.column_dimensions[get_column_letter(i)].width = 18
    ws.freeze_panes = 'A3'


# ── Sheet: Smart Pivot Tables ─────────────────────────────────────────────────
def _best_groupby_cols(df):
    """
    Return a list of up to 3 columns best suited for groupby pivots.
    Priority: fraud/label > transaction_type > explicit category cols > low-cardinality object cols.
    Skips ID-like columns and any col with too many unique values (>100).
    """
    ID_RE = re.compile(r'\b(id|uuid|key|sku|index|no\.?|#|_id)\b', re.I)
    GOOD_RE = re.compile(
        r'\b(category|type|class|label|flag|fraud|status|region|country|city|gender|segment|'
        r'channel|payment|method|department|product|brand|group|tier)\b', re.I)

    candidates = []
    for col in df.columns:
        if ID_RE.search(col):
            continue
        u = df[col].nunique()
        if u < 2 or u > 100:
            continue
        # Score: prefer named domain cols, prefer low cardinality
        priority = 0 if GOOD_RE.search(col) else 1
        candidates.append((priority, u, col))
    candidates.sort()
    return [c for _, _, c in candidates[:3]]


def build_pivot(wb, df, cat_col, num_cols):
    """
    Build meaningful pivot tables — only groups by columns that make sense
    (category, fraud flag, transaction type, etc.), never by IDs.
    Sections:
      1. One pivot per good groupby column (up to 3)
      2. Top 15 records by primary numeric column
      3. Correlation matrix with colour heatmap
    """
    ws = wb.create_sheet("🔄 Pivot Analysis")
    ws.sheet_view.showGridLines = False
    _merge_title(ws, 'A1:K1', 'PIVOT ANALYSIS — CATEGORY INTELLIGENCE', height=26)
    row = 3

    primary_num = num_cols[0] if num_cols else None
    # Pick best numeric cols (skip IDs, cap at 4)
    ID_RE = re.compile(r'\b(id|uuid|key|sku|index|no\.?|#|_id)\b', re.I)
    good_nums = [c for c in num_cols if not ID_RE.search(c)][:4]

    group_cols = _best_groupby_cols(df)
    # Always include cat_col if not already in list
    if cat_col and cat_col not in group_cols:
        group_cols.insert(0, cat_col)
    group_cols = group_cols[:3]

    # ── Pivot table per groupby column
    for gcol in group_cols:
        if not good_nums:
            break
        label = f'📊  {gcol.upper().replace("_"," ")} BREAKDOWN'
        ws.merge_cells(f'A{row}:K{row}')
        c = ws[f'A{row}']
        c.value = label
        c.font  = Font(bold=True, size=11, color=WHITE, name='Calibri')
        c.fill  = PatternFill('solid', fgColor=SLATE_BLUE)
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 20
        row += 1

        agg_dict = {nc: ['mean', 'count'] for nc in good_nums[:3]}
        try:
            pivot = df.groupby(gcol).agg(agg_dict).round(2)
            pivot.columns = [f'{col}_{agg}' for col, agg in pivot.columns]
            pivot = pivot.reset_index()
            # Sort by first count column descending
            count_col = [c for c in pivot.columns if c.endswith('_count')]
            if count_col:
                pivot = pivot.sort_values(count_col[0], ascending=False)

            hdrs = [gcol] + [
                h.replace('_mean', ' (Avg)').replace('_count', ' (Count)').replace('_', ' ')
                for h in pivot.columns[1:]
            ]
            for ci, h in enumerate(hdrs, 1):
                _hdr(ws, row, ci, h, bg=DARK_BLUE, sz=9)
            row += 1
            data_start = row
            for pi, (_, prow) in enumerate(pivot.iterrows()):
                bg = LIGHT_BLUE if pi % 2 == 0 else WHITE
                for ci, val in enumerate(prow, 1):
                    val = val if not pd.isna(val) else None
                    cell = _cell(ws, row, ci, val, bg=bg)
                    if ci > 1 and isinstance(val, (int, float)):
                        cell.number_format = '#,##0.##'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                ws.row_dimensions[row].height = 15
                row += 1

            # Data bar on first count column
            if len(pivot) > 1 and count_col:
                cc_idx = list(pivot.columns).index(count_col[0]) + 1
                ws.conditional_formatting.add(
                    f'{get_column_letter(cc_idx)}{data_start}:{get_column_letter(cc_idx)}{row-1}',
                    DataBarRule(start_type='min', end_type='max', color=MID_BLUE)
                )
            row += 2
        except Exception as e:
            ws.cell(row, 1, value=f'Pivot error for {gcol}: {e}')
            row += 2

    # ── Top 15 records by primary numeric col
    if primary_num and primary_num in df.columns:
        ws.merge_cells(f'A{row}:K{row}')
        c = ws[f'A{row}']
        c.value = f'🏆  TOP 15 RECORDS BY {primary_num.upper().replace("_"," ")}'
        c.font  = Font(bold=True, size=11, color=WHITE, name='Calibri')
        c.fill  = PatternFill('solid', fgColor=ACCENT_GREEN)
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 20
        row += 1

        # Show category + primary + up to 3 more meaningful numeric cols
        show_cols = ([cat_col] if cat_col and cat_col in df.columns else []) + \
                    [c for c in good_nums if c in df.columns][:5]
        show_cols = list(dict.fromkeys(show_cols))  # deduplicate, preserve order
        if primary_num not in show_cols:
            show_cols.insert(1, primary_num)

        top_df = df[show_cols].sort_values(primary_num, ascending=False).head(15)
        for ci, h in enumerate(show_cols, 1):
            _hdr(ws, row, ci, h.replace('_',' ').title(), bg=DARK_BLUE, sz=9)
        row += 1
        for pi, (_, prow) in enumerate(top_df.iterrows()):
            bg = LIGHT_GREEN if pi < 3 else (LIGHT_BLUE if pi % 2 == 0 else WHITE)
            for ci, val in enumerate(prow, 1):
                cell = _cell(ws, row, ci, val if not pd.isna(val) else None, bg=bg, bold=(pi < 3))
                if isinstance(val, float): cell.number_format = '#,##0.##'
            ws.row_dimensions[row].height = 15
            row += 1
        row += 2

    # ── Correlation matrix
    num_df = df[good_nums] if good_nums else df.select_dtypes(include='number')
    if len(num_df.columns) >= 2:
        ws.merge_cells(f'A{row}:K{row}')
        c = ws[f'A{row}']
        c.value = '🔗  CORRELATION MATRIX  (green = positive · red = negative · bold = strong)'
        c.font  = Font(bold=True, size=11, color=WHITE, name='Calibri')
        c.fill  = PatternFill('solid', fgColor=DARK_BLUE)
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 20
        row += 1

        corr = num_df.corr().round(3)
        _hdr(ws, row, 1, 'Column', bg=DARK_BLUE)
        for ci, col in enumerate(corr.columns, 2):
            _hdr(ws, row, ci, col.replace('_',' '), bg=DARK_BLUE, sz=9)
        row += 1
        for idx, crow in corr.iterrows():
            _hdr(ws, row, 1, str(idx).replace('_',' '), bg=SLATE_BLUE, sz=9)
            for ci, val in enumerate(crow, 2):
                cell = ws.cell(row=row, column=ci, value=round(float(val), 3))
                cell.number_format = '0.000'
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font   = Font(size=9, bold=(abs(val) > 0.7 and val != 1.0), name='Calibri')
                cell.border = _border()
                if val == 1.0:   cell.fill = PatternFill('solid', fgColor=GRAY_100)
                elif val > 0.5:  cell.fill = PatternFill('solid', fgColor='DCFCE7')
                elif val > 0.2:  cell.fill = PatternFill('solid', fgColor='F0FDF4')
                elif val < -0.5: cell.fill = PatternFill('solid', fgColor='FEF2F2')
                elif val < -0.2: cell.fill = PatternFill('solid', fgColor='FFF7ED')
                else:            cell.fill = PatternFill('solid', fgColor=WHITE)
            ws.row_dimensions[row].height = 15
            row += 1

    for i in range(1, 12):
        ws.column_dimensions[get_column_letter(i)].width = 18
    ws.freeze_panes = 'B3'

# ── Sheet: Charts ─────────────────────────────────────────────────────────────
def _place_chart(ws, buf, anchor_r, anchor_c, w=500, h=300):
    """Safely place a chart image onto a worksheet."""
    if buf is None:
        return False
    img = XLImage(buf)
    img.width  = w
    img.height = h
    ws.add_image(img, f"{get_column_letter(anchor_c)}{anchor_r}")
    return True


def build_charts(wb, df, cat_col, num_cols, fraud_col=None, amount_col=None):
    """General analytics charts: bar, pie, scatter, box — plus fraud overlay when available."""
    ws = wb.create_sheet("📊 Charts")
    ws.sheet_view.showGridLines = False
    _merge_title(ws, 'A1:T1', '📊  VISUAL ANALYTICS DASHBOARD', sz=14, height=28)

    # Skip ID-like columns for numeric axis
    ID_RE = re.compile(r'\b(id|uuid|key|sku|index|no\.?|#|_id)\b', re.I)
    clean_nums = [c for c in num_cols if not ID_RE.search(c)]

    primary_num = clean_nums[0] if clean_nums else None
    second_num  = clean_nums[1] if len(clean_nums) > 1 else None
    placed = 0

    try:
        if cat_col and primary_num:
            buf = chart_bar_top(df, cat_col, primary_num,
                                f'Avg {primary_num.replace("_"," ")} by {cat_col.replace("_"," ")} (Top 10)')
            if _place_chart(ws, buf, 2, 1):  placed += 1

        if cat_col:
            buf = chart_bar_count(df, cat_col, f'Record Count by {cat_col.replace("_"," ")}')
            if _place_chart(ws, buf, 2, 11): placed += 1

        if cat_col and primary_num:
            buf = chart_pie(df, cat_col, primary_num,
                            f'Share of {primary_num.replace("_"," ")} by {cat_col.replace("_"," ")}')
            if _place_chart(ws, buf, 20, 1): placed += 1

        if primary_num:
            buf = chart_hist(df, primary_num, f'Distribution of {primary_num.replace("_"," ")}')
            if _place_chart(ws, buf, 20, 11): placed += 1

        if cat_col and second_num:
            buf = chart_box(df, cat_col, second_num,
                            f'{second_num.replace("_"," ")} by {cat_col.replace("_"," ")}')
            if _place_chart(ws, buf, 38, 1): placed += 1

        if primary_num and second_num:
            buf = chart_scatter(df, primary_num, second_num,
                                f'{primary_num.replace("_"," ")} vs {second_num.replace("_"," ")}')
            if _place_chart(ws, buf, 38, 11): placed += 1

        # Fraud charts — placed in row 57 onwards if fraud col available
        if fraud_col:
            buf = chart_fraud_vs_legit(df, fraud_col, 'Fraud vs Legitimate Transactions')
            if _place_chart(ws, buf, 57, 1, w=480, h=300): placed += 1

        if fraud_col and amount_col:
            buf = chart_amount_by_fraud(df, fraud_col, amount_col,
                                        'Transaction Amount: Fraud vs Legitimate')
            if _place_chart(ws, buf, 57, 11, w=480, h=300): placed += 1

    except Exception as e:
        ws.cell(3, 1, value=f"Chart error: {e}")

    if placed == 0:
        ws['A3'] = 'No charts could be generated for this dataset.'


def build_distribution_charts(wb, df, cat_col, fraud_col=None, amount_col=None):
    """
    Sheet 2: Distribution-specific charts — rating buckets, discount
    buckets, products per category, and fraud charts if fraud col exists.
    """
    rating_col   = find_num_col(df, 'rating')
    discount_col = find_num_col(df, 'discount')

    if not rating_col and not discount_col and not cat_col and not fraud_col:
        return

    ws = wb.create_sheet("📉 Distribution Charts")
    ws.sheet_view.showGridLines = False
    _merge_title(ws, 'A1:T1', '📉  DISTRIBUTION ANALYSIS', sz=14, height=28)

    row_anchor = 2
    col_left   = 1
    col_right  = 11

    # ── Rating distribution (left)
    if rating_col:
        try:
            buf = chart_rating_distribution(df, rating_col, f'Rating Distribution — {rating_col}')
            _place_chart(ws, buf, row_anchor, col_left,  w=500, h=310)
        except Exception as e:
            ws.cell(row_anchor, col_left, value=f"Rating chart error: {e}")

    # ── Discount distribution (right)
    if discount_col:
        try:
            buf = chart_discount_distribution(df, discount_col, f'Discount Distribution — {discount_col}')
            _place_chart(ws, buf, row_anchor, col_right, w=500, h=310)
        except Exception as e:
            ws.cell(row_anchor, col_right, value=f"Discount chart error: {e}")

    next_row = row_anchor + 22

    # ── Fraud vs Legit count (left, row 2)
    if fraud_col:
        try:
            buf = chart_fraud_vs_legit(df, fraud_col, f'Fraud vs Legitimate — {fraud_col}')
            _place_chart(ws, buf, next_row, col_left, w=480, h=300)
        except Exception as e:
            ws.cell(next_row, col_left, value=f"Fraud chart error: {e}")

    # ── Amount by fraud (right, row 2)
    if fraud_col and amount_col:
        try:
            buf = chart_amount_by_fraud(df, fraud_col, amount_col,
                                        f'Transaction Amount: Fraud vs Legitimate')
            _place_chart(ws, buf, next_row, col_right, w=480, h=300)
        except Exception as e:
            ws.cell(next_row, col_right, value=f"Amount/fraud chart error: {e}")

    far_row = next_row + 22

    # ── Products per category (full-width bottom)
    if cat_col:
        try:
            buf = chart_category_products(df, cat_col, f'Products per Category — {cat_col}')
            _place_chart(ws, buf, far_row, col_left, w=900, h=340)
        except Exception as e:
            ws.cell(far_row, col_left, value=f"Category chart error: {e}")

    # ── Fraud by category stacked bar
    if fraud_col and cat_col:
        try:
            buf = chart_fraud_by_category(df, fraud_col, cat_col,
                                          f'Fraud vs Legitimate by {cat_col}')
            _place_chart(ws, buf, far_row, col_right, w=500, h=340)
        except Exception as e:
            ws.cell(far_row, col_right, value=f"Fraud/cat chart error: {e}")

# ── Insights Generator ────────────────────────────────────────────────────────
def generate_insights(df, cat_col, num_cols, col_info):
    """
    Produce 6–9 plain-English insights.
    Checks for domain columns (rating, discount, price, review count) first
    for specific e-commerce insights, then falls back to generic stats.
    """
    insights = []

    try:
        n_rows = len(df)
        n_num  = len(df.select_dtypes(include='number').columns)
        insights.append(
            f"📦 Dataset contains {n_rows:,} rows and {len(df.columns)} columns "
            f"with {n_num} numeric columns detected after cleaning."
        )

        # ── Missing data
        total_missing = int(df.isna().sum().sum())
        if total_missing > 0:
            worst     = df.isna().sum().idxmax()
            worst_pct = df[worst].isna().mean() * 100
            insights.append(
                f"⚠️  {total_missing:,} missing values across the dataset. "
                f"'{worst}' has the most missing data ({worst_pct:.1f}%)."
            )
        else:
            insights.append("✅ No missing values — dataset is complete.")

        # ── Rating insights
        rating_col = find_num_col(df, 'rating')
        if rating_col:
            data = df[rating_col].dropna()
            mean_r = data.mean()
            bins   = [0, 3.5, 4.0, 4.5, 5.01]
            labels = ['Below 3.5', '3.5–4.0', '4.0–4.5', '4.5–5.0']
            bucket_counts = pd.cut(data, bins=bins, labels=labels, right=False).value_counts()
            dominant_bucket = bucket_counts.idxmax()
            dominant_pct    = bucket_counts.max() / len(data) * 100
            insights.append(
                f"⭐ Most products are rated in the {dominant_bucket} range "
                f"({dominant_pct:.0f}% of products). Overall average rating: {mean_r:.2f}."
            )

        # ── Discount insights
        discount_col = find_num_col(df, 'discount')
        if discount_col:
            data  = df[discount_col].dropna()
            avg_d = data.mean()
            high_disc_pct = (data >= 50).sum() / len(data) * 100
            insights.append(
                f"💸 Average discount is {avg_d:.1f}%. "
                f"{high_disc_pct:.0f}% of products have a discount of 50% or more."
            )

        # ── Category insights
        if cat_col and cat_col in df.columns:
            # Largest category (use last segment if pipe-separated)
            top_level = df[cat_col].astype(str).str.split('|').str[-1].str.strip()
            biggest_cat   = top_level.value_counts().idxmax()
            biggest_count = top_level.value_counts().max()
            biggest_pct   = biggest_count / n_rows * 100
            insights.append(
                f"📂 '{biggest_cat}' is the largest category with {biggest_count:,} products "
                f"({biggest_pct:.1f}% of the dataset)."
            )

            # Best-rated category
            if rating_col and rating_col in df.columns:
                cat_ratings = df.groupby(cat_col)[rating_col].mean()
                best_cat    = cat_ratings.idxmax()
                best_val    = cat_ratings.max()
                insights.append(
                    f"🏆 Highest average rating by category: '{best_cat}' with {best_val:.2f} stars."
                )

        # ── Review count insights
        review_col = find_num_col(df, 'rating_count', 'review_count', 'no_of_ratings', 'reviews')
        if review_col and rating_col:
            high_review_threshold = df[review_col].quantile(0.75)
            high_review_df = df[df[review_col] >= high_review_threshold]
            avg_rating_high = high_review_df[rating_col].mean() if not high_review_df.empty else None
            if avg_rating_high is not None:
                insights.append(
                    f"🔗 Products with high review counts (top 25%) have an average rating "
                    f"of {avg_rating_high:.2f} — "
                    f"{'above' if avg_rating_high > df[rating_col].mean() else 'below'} the overall average."
                )

        # ── Fraud-specific insights
        fraud_col_i, is_binary = detect_fraud_col(df)
        if fraud_col_i and is_binary:
            def to_01(v):
                return 1 if str(v).lower() in ('1','true','yes','fraud') else 0
            fraud_series = df[fraud_col_i].apply(to_01)
            n_fraud_i = int(fraud_series.sum())
            n_total_i = len(fraud_series)
            fraud_pct_i = n_fraud_i / n_total_i * 100
            insights.append(
                f"🚨 Fraud rate: {n_fraud_i:,} of {n_total_i:,} transactions are fraudulent "
                f"({fraud_pct_i:.2f}%)."
            )
            if cat_col and cat_col in df.columns:
                grp = df.copy()
                grp['_f'] = fraud_series
                rates = grp.groupby(cat_col)['_f'].mean() * 100
                high_risk = rates.idxmax()
                high_rate = rates.max()
                insights.append(
                    f"⚠️  Highest fraud rate by {cat_col}: '{high_risk}' ({high_rate:.1f}% fraud)."
                )
            amount_col_i = find_num_col(df, 'amount', 'transaction_amount', 'value', 'sum')
            if amount_col_i:
                df2 = df.copy(); df2['_f'] = fraud_series
                avg_fraud = df2[df2['_f']==1][amount_col_i].mean()
                avg_legit = df2[df2['_f']==0][amount_col_i].mean()
                if pd.notna(avg_fraud) and pd.notna(avg_legit):
                    ratio = avg_fraud / avg_legit if avg_legit != 0 else 0
                    insights.append(
                        f"💰 Average fraudulent transaction (${avg_fraud:,.2f}) is "
                        f"{ratio:.1f}× the average legitimate transaction (${avg_legit:,.2f})."
                    )
            primary = num_cols[0]
            insights.append(
                f"📈 {primary}: min={df[primary].min():,.2f}, "
                f"max={df[primary].max():,.2f}, mean={df[primary].mean():,.2f}."
            )

        # ── Correlation highlight
        if len(num_cols) >= 2:
            corr = df[num_cols].corr()
            pairs = [(abs(corr.loc[a, b]), corr.loc[a, b], a, b)
                     for i, a in enumerate(num_cols)
                     for j, b in enumerate(num_cols) if i < j]
            if pairs:
                _, val, ca, cb = max(pairs)
                direction = 'positively' if val > 0 else 'negatively'
                if abs(val) >= 0.3:
                    insights.append(
                        f"🔗 '{ca}' and '{cb}' are {direction} correlated (r = {val:.2f})."
                    )

    except Exception as e:
        insights.append(f"(Insight generation partial error: {e})")

    return insights

# ── Summary JSON for React dashboard ─────────────────────────────────────────
def compute_summary_json(df, cat_col, num_col):
    try:
        if df[num_col].dtype == object:
            df[num_col] = pd.to_numeric(df[num_col], errors='coerce')
        if df[num_col].isna().all():
            df['_count'] = 1; num_col = '_count'

        pivot = df.groupby(cat_col)[num_col].mean().sort_values(ascending=False).head(10)
        trend = df.groupby(cat_col)[num_col].mean()

        def sf(v):
            try: f=float(v); return round(f,2) if abs(f)<1e15 else 0
            except: return 0

        return {
            "kpis": {
                "total":   sf(df[num_col].sum()),
                "mean":    sf(df[num_col].mean()),
                "max":     sf(df[num_col].max()),
                "min":     sf(df[num_col].min()),
                "rows":    int(len(df)),
                "cols":    int(len(df.columns)),
                "missing": round(float(df.isnull().mean().mean())*100, 1),
            },
            "bar_chart":  [{"name":str(k)[:22],"value":sf(v)} for k,v in pivot.items()],
            "pie_chart":  [{"name":str(k)[:22],"value":sf(v)} for k,v in pivot.items()],
            "line_chart": [{"name":str(k)[:22],"value":sf(v)} for k,v in trend.items()],
            "cat_col": cat_col, "num_col": num_col,
        }
    except: return None

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 3:
        print(json.dumps({"error":"Usage: report_engine.py <input> <output>"})); sys.exit(1)

    input_path  = sys.argv[1]
    output_path = sys.argv[2]

    try:
        # 1. Load
        emit("load", 5, "Loading file...")
        df_raw = load_data(input_path)
        original_rows = len(df_raw)
        original_cols = len(df_raw.columns)

        # 2. Clean
        emit("clean", 15, "Cleaning data & detecting numeric columns...")
        df, cleaning_log = clean_data(df_raw)
        duplicate_rows = original_rows - len(df)

        # 3. Analyze
        emit("analyze", 28, "Analyzing columns...")
        col_info = analyze_columns(df)
        cat_col, num_col = pick_cols(df)
        num_cols = df.select_dtypes(include='number').columns.tolist()

        # Detect domain-specific columns
        name_col         = find_col(df, 'product_name', 'product_title', 'name', 'title', 'item')
        rating_col       = find_num_col(df, 'rating')
        rating_count_col = find_num_col(df, 'rating_count', 'review_count', 'no_of_ratings', 'reviews')
        discount_col     = find_num_col(df, 'discount')
        price_col        = find_num_col(df, 'discounted_price', 'sale_price', 'price', 'cost')
        # Fraud / transaction dataset columns
        fraud_col, _fraud_binary = detect_fraud_col(df)
        amount_col   = find_num_col(df, 'amount', 'transaction_amount', 'trans_amount', 'value', 'sum')
        tx_type_col  = find_col(df, 'type', 'transaction_type', 'trans_type', 'payment_type', 'method')

        stats_dict = {
            "rows":           len(df),
            "cols":           len(df.columns),
            "numeric_cols":   len(num_cols),
            "missing_pct":    float(df.isna().mean().mean() * 100),
            "duplicate_rows": duplicate_rows,
        }

        # 4. Insights
        emit("insights", 35, "Generating insights...")
        insights = generate_insights(df, cat_col, num_cols, col_info)

        # 5. Build workbook
        emit("sheets", 45, "Building report sheets...")
        wb = Workbook()

        # Cover / Summary
        try:
            build_cover(wb, col_info, stats_dict, cleaning_log, insights)
        except Exception as e:
            print(json.dumps({"warning": f"Cover: {e}"}), flush=True)

        # Cleaned data
        emit("raw", 55, "Writing cleaned data...")
        try:
            build_raw(wb, df, cleaning_log)
        except Exception as e:
            print(json.dumps({"warning": f"Raw: {e}"}), flush=True)

        # Statistics
        emit("stats", 63, "Computing statistics...")
        try:
            build_stats(wb, df)
        except Exception as e:
            print(json.dumps({"warning": f"Stats: {e}"}), flush=True)

        # Missing value sheet
        emit("missing", 68, "Analyzing missing values...")
        try:
            build_missing(wb, df)
        except Exception as e:
            print(json.dumps({"warning": f"Missing: {e}"}), flush=True)

        # Pivot
        emit("pivot", 75, "Building pivot tables...")
        try:
            build_pivot(wb, df, cat_col, num_cols)
        except Exception as e:
            print(json.dumps({"warning": f"Pivot: {e}"}), flush=True)
            ws_p = wb.create_sheet("🔄 Pivot Analysis")
            ws_p["A1"] = f"Pivot unavailable: {e}"

        # Charts
        emit("charts", 82, "Generating analytics charts...")
        try:
            build_charts(wb, df, cat_col, num_cols, fraud_col=fraud_col, amount_col=amount_col)
        except Exception as e:
            print(json.dumps({"warning": f"Charts: {e}"}), flush=True)
            wb.create_sheet("📊 Charts")["A1"] = f"Charts unavailable: {e}"

        # Distribution charts (rating / discount / category / fraud)
        emit("distcharts", 87, "Generating distribution charts...")
        try:
            build_distribution_charts(wb, df, cat_col, fraud_col, amount_col)
        except Exception as e:
            print(json.dumps({"warning": f"Distribution charts: {e}"}), flush=True)

        # Category analysis
        emit("category", 89, "Building category analysis...")
        try:
            build_category_analysis(wb, df, cat_col, rating_col, discount_col, price_col)
        except Exception as e:
            print(json.dumps({"warning": f"Category analysis: {e}"}), flush=True)

        # Top products
        emit("topproducts", 91, "Building top products sheet...")
        try:
            build_top_products(wb, df, name_col, rating_col, rating_count_col, discount_col, cat_col)
        except Exception as e:
            print(json.dumps({"warning": f"Top products: {e}"}), flush=True)

        # Fraud analysis (only if fraud column found)
        if fraud_col:
            emit("fraud", 93, "Building fraud pattern analysis...")
            try:
                build_fraud_analysis(wb, df, fraud_col, cat_col, amount_col, tx_type_col)
            except Exception as e:
                print(json.dumps({"warning": f"Fraud analysis: {e}"}), flush=True)

        # Outlier detection
        emit("outliers", 94, "Running outlier detection...")
        try:
            build_outlier_analysis(wb, df, num_cols[:8])  # cap at 8 numeric cols
        except Exception as e:
            print(json.dumps({"warning": f"Outlier detection: {e}"}), flush=True)

        # Save
        emit("save", 97, "Saving report...")
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)

        # Summary for React
        summary = compute_summary_json(df, cat_col, num_col)

        emit("done", 100, "Report complete!", {
            "rows":       len(df),
            "cols":       len(df.columns),
            "cat_column": cat_col,
            "num_column": num_col,
            "output":     output_path,
            "cols_info":  col_info,
            "summary":    summary,
        })

    except Exception as e:
        print(json.dumps({"error": str(e), "trace": traceback.format_exc()}), flush=True)
        sys.exit(1)

if __name__ == "__main__":
    main()
